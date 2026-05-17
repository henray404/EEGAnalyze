import io
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Any, List, Dict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter()

_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_HEADER_FILL = PatternFill("solid", fgColor="122056")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_ALT_FILL    = PatternFill("solid", fgColor="F5F6FE")
_META_COLS   = {"filename", "category", "subject", "task", "channel", "subband", "time", "scenario"}


class SheetSpec(BaseModel):
    name: str
    records: List[Dict[str, Any]]


class ExcelRequest(BaseModel):
    sheets: List[SheetSpec]
    filename: str = "batch_results"


def _write_sheet(ws, records: list[dict]):
    if not records:
        ws.cell(1, 1, "Tidak ada data").font = Font(italic=True)
        return

    headers = list(records[0].keys())
    col_widths = {h: max(len(str(h)), 8) for h in headers}

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _THIN

    for ri, rec in enumerate(records, 2):
        fill = _ALT_FILL if ri % 2 == 0 else None
        for ci, h in enumerate(headers, 1):
            v = rec.get(h)
            if isinstance(v, float):
                v = round(v, 8)
            cell = ws.cell(ri, ci, v)
            cell.border = _THIN
            if fill:
                cell.fill = fill
            if h not in _META_COLS and isinstance(v, (int, float)):
                cell.number_format = "0.000000E+00"
                cell.alignment = Alignment(horizontal="right")
            col_widths[h] = max(col_widths[h], len(str(v or "")) + 2)

    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = min(col_widths[h], 30)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


@router.post("/excel")
async def export_excel(req: ExcelRequest):
    """Generate Excel dari list of sheets berisi records."""
    if not req.sheets:
        raise HTTPException(status_code=400, detail="Minimal satu sheet diperlukan")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for spec in req.sheets:
        safe_name = spec.name[:31].replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "").replace("[", "").replace("]", "")
        ws = wb.create_sheet(safe_name or "Sheet")
        _write_sheet(ws, spec.records)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_file = req.filename.replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_file}.xlsx"'},
    )
